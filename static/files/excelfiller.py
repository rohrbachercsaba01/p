from openpyxl import load_workbook
from datetime import datetime, timedelta

def fillFullAttendance(attendance_data, holidays_data):
    workHoursByMonth = {
    '01' : '1',
    '02' : '2',
    '03' : '3',
    '04' : '4',
    '05' : '5',
    '06' : '6',
    '07' : '7',
    '08' : '8',
    '09' : '9',
    '10' : '10',
    '11' : '11',
    '12' : '12',
    }
    
    def get_dates_between(start_date, end_date):
        # Convert input strings to datetime objects
        #start_date = datetime.strptime(str(start_date), '%Y-%m-%d')
        #end_date = datetime.strptime(str(end_date), '%Y-%m-%d')

        # Initialize an empty list to store the dates
        dates_list = []

        # Iterate through the range of dates and append them to the list
        current_date = start_date
        while current_date <= end_date:
            dates_list.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

        return dates_list
    
    def is_weekend(date_str):
        date = datetime.strptime(str(date_str), '%Y-%m-%d')
        if date.weekday() in [5, 6]:  # Saturday is 5, Sunday is 6
            return "yes"
        else:
            return "no"


    def calculate_time_difference(start_hour, end_hour):
        start_time = datetime.strptime(start_hour, "%H:%M:%S")
        end_time = datetime.strptime(end_hour, "%H:%M:%S")
        time_difference = end_time - start_time
        return time_difference 

    def calculate_endtime(endTime):
        end_time_dt = datetime.strptime(endTime, '%H:%M:%S')

        if calculate_time_difference(startTime, endTime) >= timedelta(hours=6, minutes=30):
            new_time_dt = end_time_dt - timedelta(minutes=30)
            new_hour = new_time_dt.strftime('%H')
            new_minute = new_time_dt.strftime('%M')
            return new_hour, new_minute
        else: 
            new_time_dt = end_time_dt
            new_hour = new_time_dt.strftime('%H')
            new_minute = new_time_dt.strftime('%M')
            return new_hour, new_minute
    
    wb = load_workbook('static/files/jelenleti.xlsx')

    wsÁ = wb['Állomány']

    ws = wb['Alap']

    current_month = datetime.now().strftime('%m')
    wsÁ.cell(1, 18).value = str(current_month)
            
    # Find the starting column for dates
    start_column = 3  # Assuming the first column is for names

    # Fill in the attendance data
    for name, entries in attendance_data.items():
        # Find the row corresponding to the worker's name
        row = 1
        while row <= ws.max_row:
            if ws.cell(row, 1).value == name:
                break
            row += 1
        else:
            print(f"Worker {name} not found in the template.")
            continue

        # Fill in the start and end hours for each date
        for entry in entries:
            date = datetime.strptime(entry['date'], "%Y-%m-%d")
            startTime = entry['startHour']
            start_hour, start_minute, start_second = startTime.split(":")
            endTime = entry['endHour']

            # Find the column corresponding to the date
            column = start_column + (date.day - 1) * 5  # Adjust the column index to write start and end hours side by side

            # Write the start and end hours
            ws.cell(row, column).value = int(start_hour)
            ws.cell(row, column + 1).value = int(start_minute)
            ws.cell(row, column + 2).value = calculate_endtime(endTime)[0]
            ws.cell(row, column + 3).value = calculate_endtime(endTime)[1]

    for name, entries in holidays_data.items():
        # Find the row corresponding to the worker's name
        row = 1
        while row <= ws.max_row:
            if ws.cell(row, 1).value == name:
                break
            row += 1
        else:
            print(f"Worker {name} not found in the template.")
            continue

        for entry in entries:
            if entry['szabadsagType'] == 'Szabadság':
                szabiT = 'sz'
            elif entry['szabadsagType'] == 'Betegszabi':
                szabiT = 'b'
            Sdate = datetime.strptime(entry['startDate'], "%Y-%m-%d")
            Edate = datetime.strptime(entry['endDate'], "%Y-%m-%d")
            datesBetween = get_dates_between(Sdate, Edate)
            for date in datesBetween:
                if is_weekend(date) == "no":
                    date = datetime.strptime(date, "%Y-%m-%d")
                    # Find the column corresponding to the date
                    column = start_column + (date.day - 1) * 5  # Adjust the column index to write start and end hours side by side

                    # Write the start and end hours
                    ws.cell(row, column).value = szabiT
                elif is_weekend(date) == "yes":
                    pass


    # Save the updated workbook
    wb.save('jelenleti.xlsx')
    return 'jelenleti.xlsx'
